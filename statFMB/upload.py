from datetime import date, datetime

import xlrd
from openpyxl import load_workbook, worksheet
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

from .statFMB import db, Report, current_user
from .models import *
from .utils import clear_str, represents_int

#TODO: create a pending entrances, mainly for email script but also
#      so we don't loose invalid entrances

#returns a Dict with results from the update, indexed by file
def update_database(new_files):

    upload_results = {}
    failed_uploads = {}
    for new_file in new_files:

        report,entrance_obj_list,error_list,error_msg = upload_file(new_file)
        if report != "error":
            upload_results[new_file.filename] = [report,
                                                 entrance_obj_list,
                                                 error_list]
        else:
            failed_uploads[new_file.filename] = error_msg

    return upload_results, failed_uploads


def upload_file(new_file):
    print("inserting: {} in database.".format(new_file.filename))

    ##load uploaded file and sheets
    try:
        if new_file.filename[-4:] == ".xls":
            wb = open_xls_as_xlsx(new_file)
        else:
            wb = load_workbook(new_file, read_only = True, data_only = True)
    except Exception as err:
        print (err)
        return "error",[],[],"Ficheiro inválido"

    try:
        statSheet = wb.get_sheet_by_name("Estatística")
        regSheet = wb.get_sheet_by_name("Folha de Registo")
    except Exception as err:
        print(err)
        return "error",[],[],"Nome das folhas inválido"

    ##date
    try:
        date = format_date(regSheet['C6'].value)
        print("==> date: {}".format(date))
    except Exception as err:
        print (err)
        return "error",[],[],"Data inválida"

    ##gate
    try:
        gate_str = regSheet['A4'].value.rsplit(' ',1)[-1].capitalize()
        print ("==> Gate: " + gate_str)
        gate = Gate.get_gate(gate_str)
    except Exception as err:
        print (err)
        return "error",[],[],"Porta inválida"

    ##hours
    try:
        start_time_str, end_time_str = get_hours(regSheet)
        start_time_str = date.strftime("%d-%m-%Y-") + start_time_str
        end_time_str = date.strftime("%d-%m-%Y-") + end_time_str
        start_time = datetime.strptime(start_time_str,"%d-%m-%Y-%H-%M")
        end_time = datetime.strptime(end_time_str,"%d-%m-%Y-%H-%M")

        if start_time >= end_time:
            raise Exception('start_time >= end_time')

        print ("==> Hours: {} -> {}".format(start_time,end_time))
    except Exception as err:
        print(err)
        return "error",[],[],"Hora inválida"

    ##vehicles
    try:
        vehicles = get_vehicles(regSheet)
        print ("==> Vehicles: " + str(vehicles['total']))
    except Exception as err:
        print (err)
        return "error",[],[],"Numero de veículos inválido"

    #pawns
    try:
        pawns = get_pawns(statSheet)
        print ("==> Pawns: " + str(pawns))
    except Exception as err:
        print (err)
        return "error",[],[],"Numero de pessoas a pé inválido"

    #bicicles
    try:
        bicicles = get_bicicles(statSheet)
        print ("==> Bicicles: " + str(bicicles))
    except Exception as err:
        print (err)
        return "error",[],[],"Numero de bicicletas inválido"

    #NOTE: vehicles dict cannot be passed as a parameter to the
    #      Report __init__(), this leaves sqlalchemy in a bad state that
    #      generates a bug in reading from the db later on
    report = Report(date = date,
                    start_time = start_time,
                    end_time = end_time,
                    validated = False,
                    vehicles = vehicles["total"],
                    bikes = vehicles["bikes"],
                    lightduty = vehicles["lightduty"],
                    lightdutyXL = vehicles["lightdutyXL"],
                    caravans = vehicles["caravans"],
                    busses = vehicles["busses"],
                    pawns = pawns,
                    bicicles = bicicles,
                    gate = gate,
                    user = current_user,
    )
    print("==> Report Created!")

    print("==> Instanciating entrances!")
    #entrances
    entrance_obj_list, error_list = get_entrance_list(statSheet,report)
    print ("==> Entrances registered: {}".format(len(entrance_obj_list)))
    print ("==> Errors found: {}".format(len(error_list)))

    #prevent the case when the user forgets to register the "Folha de Registo"
    if (vehicles["total"] < (len(entrance_obj_list) + len(error_list))):
        return "error",[],[],"Numero de veículos da folha de estatística é maior que o numero de veículos na folha de registo"

    #passengers
    passengers = 0
    for entrance in entrance_obj_list:
        passengers += entrance.passengers

    new_file.close()
    print ("All instances Created, updating database!")

    try:
        report_db_action = Report.get_db_action(date,gate, start_time,end_time)
        if report_db_action == "replace":
            report_old = Report.get_report(date,start_time,end_time,gate)
            entrances_old = Entrance.get_entrances_of_report(report_old)

            print("Report already exists in the database, replacing!")
            print("==> {} entrances deleted!".format(len(entrances_old)))
            print("==> Report {} deleted!".format(report_old.id))

            for entrance in entrances_old:
                db.session.delete(entrance)
            db.session.delete(report_old)
        elif report_db_action == "invalid":
            raise Exception(
                "==> Error, new report is within the time range of another report")
    except Exception as err:
        print (err)
        return ("error",[],[],
                "Já existe um registo na base de dados durante \
                este período ({} -> {})".format(start_time, end_time))

    db.session.add(report)
    for entrance in entrance_obj_list:
        db.session.add(entrance)
    print("database updated for: {} with report id: {}"
          .format(new_file.filename,report.id))

    if report_db_action == "valid":
        description = "carregou {}".format(report)
    elif report_db_action == "replace":
        description = "carregou {} substituindo {}".format(report,report_old)

    log = Log(description = description,user = current_user)
    db.session.add(log)
    db.session.commit()
    error_message = ""

    return report, entrance_obj_list, error_list, error_message


def save_corrections(saved_corrections):

    for report_id in saved_corrections:
        report = Report.get_report_by_id(report_id)
        for entry in saved_corrections[report_id]:
            vt_alias_list = Vehicle_type_alias.get_alias_list()
            c_alias_list = Country_alias.get_alias_list()
            m_alias_list = Municipality_alias.get_alias_list()

            vehicle_type_obj = Vehicle_type.get_vehicle_type(
                entry["vehicle_type"])
            country_obj = Country.get_country(entry["country"])
            municipality_obj = Municipality.get_municipality(
                entry["municipality"])

            entrance_obj =Entrance(report = report,
                                   vehicle_type = vehicle_type_obj,
                                   passengers=int(float(entry["passengers"])),
                                   country = country_obj,
                                   municipality = municipality_obj)
            db.session.add(entrance_obj)

            #update vehicle type alias table
            if (vehicle_type_obj.vehicle_type!= entry["vehicle_type_failed"]):
                if (entry["vehicle_type_failed"] not in vt_alias_list):
                    vehicle_type_alias_obj = Vehicle_type_alias(
                        alias = entry["vehicle_type_failed"],
                        vehicle_type = vehicle_type_obj)
                    vt_alias_list.append(vehicle_type_alias_obj)
                    db.session.add(vehicle_type_alias_obj)

            #update country alias table
            if (entry["country_failed"] != "N/A"
                and country_obj.country != entry["country_failed"]):
                if (entry["country_failed"] not in c_alias_list):
                    country_alias_obj = Country_alias(
                        alias = entry["country_failed"],
                        country = country_obj)
                    c_alias_list.append(country_alias_obj)
                    db.session.add(country_alias_obj)

            #update municipality alias table
            if (entry["municipality_failed"] != "N/A"
                and municipality_obj.municipality
                != entry["municipality_failed"]):
                if (entry["municipality_failed"] not in m_alias_list):
                    municipality_alias_obj = Municipality_alias(
                        alias = entry["municipality_failed"],
                        municipality = municipality_obj)
                    m_alias_list.append(municipality_alias_obj)
                    db.session.add(municipality_alias_obj)

    db.session.commit()

    return


#returns a tuple (entrance_obj_list, error_list)
def get_entrance_list(sheet,report):
    #get list of payed entrances
    col_number = 1
    upper_bound = "Dados visitantes veículos pagantes"
    lower_bound = "Dados visitantes veículos não pagantes " + \
                  "(hóspedes, reuniões, outros)"
    interval = find_row_interval(sheet, col_number,
                                 upper_bound, lower_bound)

    vehicle_types_list = Vehicle_type.get_vehicle_types_list()
    countries_list = Country.get_countries_list()
    municipalities_list = Municipality.get_municipalities_list()


    vt_alias_dict = Vehicle_type_alias.get_dict()
    c_alias_dict = Country_alias.get_dict()
    m_alias_dict = Municipality_alias.get_dict()

    entrance_obj_list = []
    error_list = []

    #NOTE: interval[0]+2 and interval[1]-1 are related to the way the file is
    #      formated
    for row in sheet.iter_rows(min_row=interval[0]+2,
                               max_row=interval[1]-1,
                               max_col=19):
        if row[0].value:
            ## Vehicle_type validation
            vehicle_type = clear_str(word = row[0].value.capitalize(),
                                     list_to_check = vehicle_types_list,
                                     dict_to_double_check = vt_alias_dict)
            if vehicle_type == "invalid":
                vehicle_type = row[0].value.capitalize()

            ##passengers validation
            if represents_int(row[6].value): passengers = row[6].value
            else: passengers = 1

            ## Municipality validation
            if row[18].value:
                municipality = clear_str(row[18].value.capitalize(),
                                         list_to_check = municipalities_list,
                                         dict_to_double_check = m_alias_dict)
                if municipality == "invalid":
                    municipality = row[18].value.capitalize()
            else:
                #NOTE: this sets de default for ""
                municipality = "N/A"

            #Country vaidation
            #if municipality is set, country must be "Portugal"
            if municipality != "N/A":
                country = "Portugal"
            elif row[12].value:
                country = clear_str(row[12].value.capitalize(),
                                    list_to_check = countries_list,
                                    dict_to_double_check = c_alias_dict)
                if country == "invalid":
                    country = row[12].value
            else:
                #NOTE: this sets de default for ""
                country = "N/A"

            #test if everything is valid
            if (vehicle_type in vehicle_types_list
                and (country in countries_list
                     and country != "N/A")
                and (municipality in municipalities_list
                     and municipality != "N/A"
                     or (municipality == "N/A"
                         and country != "Portugal"))):
                entrance = Entrance(passengers = passengers,
                                    report = report,
                                    vehicle_type = Vehicle_type
                                    .get_vehicle_type(vehicle_type),
                                    country=Country.get_country(country),
                                    municipality = Municipality
                                    .get_municipality(municipality),
                )

                entrance_obj_list.append(entrance)
            else:
                error_list.append([vehicle_type,
                                   passengers,
                                   country,
                                   municipality])

    return entrance_obj_list, error_list


def get_hours(sheet):
    start_hour = str(sheet['R5'].value)
    start_minute = str(sheet['T5'].value)
    end_hour = str(sheet['V5'].value)
    end_minute = str(sheet['X5'].value)

    start_time_str = start_hour + "-" + start_minute
    end_time_str = end_hour + "-" + end_minute

    return start_time_str, end_time_str


def get_vehicles(sheet):
    vehicles = {"total": 0,
                "bikes": 0,
                "lightduty": 0,
                "lightdutyXL": 0,
                "caravans": 0,
                "busses": 0}

    if sheet['T11'].value : vehicles['bikes'] += sheet['T11'].value
    if sheet['T12'].value : vehicles['bikes'] += sheet['T12'].value
    if sheet['T13'].value : vehicles['lightduty'] += sheet['T13'].value
    if sheet['T14'].value : vehicles['lightduty'] += sheet['T14'].value
    if sheet['T15'].value : vehicles['lightdutyXL'] += sheet['T15'].value
    if sheet['T16'].value : vehicles['caravans'] += sheet['T16'].value
    if sheet['T17'].value : vehicles['busses'] += sheet['T17'].value
    if sheet['T18'].value : vehicles['busses'] += sheet['T18'].value

    vehicles['total'] = (vehicles['bikes'] +
                         vehicles['lightduty'] +
                         vehicles['lightdutyXL'] +
                         vehicles['caravans'] +
                         vehicles['busses']
    )

    return vehicles


def get_pawns(sheet):
    col_number = 1
    col_leter = 'A'
    lower_bound = "N.º entradas a pé"
    interval = find_row_interval(sheet,
                                 col_number,
                                 lower_delimiter = lower_bound)

    pawn_list = str_to_int_list(sheet[col_leter + str(interval[0]+1) :
                                      col_leter + str(interval[1])])

    return sum(pawn_list)


def get_bicicles(sheet):
    col_number = 9
    col_leter = 'I'
    lower_bound = "N.º entradas de bicicleta"
    interval = find_row_interval(sheet,
                                 col_number,
                                 lower_delimiter = lower_bound)

    bicicle_list = str_to_int_list(sheet[col_leter + str(interval[0]+1) :
                                         col_leter + str(interval[1])])
    return sum(bicicle_list)


#returns the row interval between lower_delimiter and upper_delimiter
def find_row_interval(sheet, column = 1,
                      lower_delimiter = None, upper_delimiter = None):
    upper_bound = 0
    lower_bound = 0

    for row in sheet.iter_rows(min_col = column-1,max_col = column):
        for cell in row:
            if lower_delimiter:
                if cell.value == lower_delimiter:
                    lower_bound = cell.row
            else:
                upper_delimiter = 0

            if upper_delimiter:
                if cell.value == upper_delimiter:
                    upper_bound = cell.row
            else:
                upper_delimiter = sheet.max_row

    return (lower_bound,upper_bound)


#expects a string "dd-mm-yyyy" or "dd/mm/yyyy" and formats to iso
def format_date(d):
    if len(d.split('/')) == 3:
        d_list = d.split('/')
    elif len(d.split('-') == 3):
        d_list = d.split('-')
    else:
        #return error
        print("error spliting date")

    formated_date = date(int(d_list[2]),int(d_list[1]),int(d_list[0]))

    return formated_date


def str_to_int_number(n):
    n_str = ''.join(c for c in n if c.isdigit())
    if n_str == "":
        return 0
    else:
        return int(n_str)


def str_to_int_list(l):
    new_list = []
    for element in l:
        if element[0].value != None:
            new_list.append(str_to_int_number(str(element[0].value)))
    return new_list


def open_xls_as_xlsx(filename):
    # open using xlrd
    book = xlrd.open_workbook(file_contents = filename.read())

    regSheet_xls_nrows, regSheet_xls_ncols = 0, 0
    while regSheet_xls_nrows * regSheet_xls_ncols == 0:
        regSheet_xls = book.sheet_by_name("Folha de Registo")
        regSheet_xls_nrows = regSheet_xls.nrows + 1
        regSheet_xls_ncols = regSheet_xls.ncols + 1

    statSheet_xls_nrows, statSheet_xls_ncols = 0, 0
    while statSheet_xls_nrows * statSheet_xls_ncols == 0:
        statSheet_xls = book.sheet_by_name("Estatística")
        statSheet_xls_nrows = statSheet_xls.nrows + 1
        statSheet_xls_ncols = statSheet_xls.ncols + 1

    #obsSheet_xls_nrows, obsSheet_xls_ncols = 0, 0
    #while obsSheet_xls_nrows * obsSheet_xls_ncols == 0:
    #    obsSheet_xls = book.sheet_by_name("Observações")
    #    obsSheet_xls_nrows = obsSheet_xls.nrows + 1
    #    obsSheet_xls_ncols = obsSheet_xls.ncols + 1

    # prepare a xlsx sheet
    book_xlsx = Workbook()
    regSheet = book_xlsx.create_sheet("Folha de Registo")
    statSheet = book_xlsx.create_sheet("Estatística")
    #obsSheet = book_xlsx.create_sheet("Observações")

    for row in range(1, regSheet_xls_nrows):
        for col in range(1, regSheet_xls_ncols):
            regSheet.cell(row=row, column=col).value = regSheet_xls.cell_value(
                row-1,
                col-1)

    for row in range(1, statSheet_xls_nrows):
        for col in range(1, statSheet_xls_ncols):
            statSheet.cell(row=row, column=col).value=statSheet_xls.cell_value(
                row-1,
                col-1)

    #for row in range(1, obsSheet_xls_nrows):
    #    for col in range(1, obsSheet_xls_ncols):
    #        obsSheet.cell(row=row, column=col).value =obsSheet_xls.cell_value(
    #            row-1,
    #            col-1)

    #converting date types
    # 3 means 'xldate' , 1 means 'text'
    if regSheet_xls.cell(5,2).ctype == 3:
        date_field = statSheet_xls.cell(5,2).value
        py_date = datetime(*xlrd.xldate_as_tuple(date_field,book.datemode))

        #this if solves a werid behaviour of xldate_as_tuple()
        #that swaps the day and month when day <= 12
        #NOTE:should take a look at xlrd.xldate_as_datetime
        #   https://github.com/python-excel/xlrd/blob/master/xlrd/xldate.py
        if py_date.day > 12:
            regSheet.cell(row=6, column=3).value = "{}/{}/{}".format(
                py_date.day,
                py_date.month,
                py_date.year)
        else:
            regSheet.cell(row=6, column=3).value = "{}/{}/{}".format(
                py_date.month,
                py_date.day,
                py_date.year)


    #opens the xlsx file for debug only
    #path = "/home/enemabandit/lab/FMB/statFMB/statFMB/uploads/TESTEXLS.xlsx"
    #book_xlsx.save(path)
    #import subprocess
    #process = subprocess.call("localc " + path ,shell=True)

    return book_xlsx
